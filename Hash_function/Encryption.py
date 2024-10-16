import hashlib
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl

def SelectInputFile():
    filePath = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
    if filePath:
        EntryInputFile.delete(0, tk.END)
        EntryInputFile.insert(0, filePath)

def SelectOutputFile():
    filePath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if filePath:
        EntryOutputFile.delete(0, tk.END)
        EntryOutputFile.insert(0, filePath)

def HashPhoneNumbers(HashFunction):
    inputFile = EntryInputFile.get()
    outputFile = EntryOutputFile.get()
    if not inputFile or not outputFile:
        messagebox.showerror("Error", "Please select an input file and a location to save the output.")
        return

    try:
        with open(inputFile, 'r') as file:
            phoneNumbers = file.readlines()

        hashedPhoneNumbers = []
        for phoneNumber in phoneNumbers:
            phoneNumber = phoneNumber.strip()
            hashObject = HashFunction(phoneNumber.encode())
            hashedPhoneNumbers.append(hashObject.hexdigest())

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Hashes"

        worksheet.cell(row=1, column=1, value="Phone Number Hash")

        for index, hashedPhoneNumber in enumerate(hashedPhoneNumbers, start=2):
            worksheet.cell(row=index, column=1, value=hashedPhoneNumber)

        workbook.save(outputFile)
        messagebox.showinfo("Success", f"Hashing completed! Hashes saved to {outputFile}")

    except Exception as exception:
        messagebox.showerror("Error", f"An error occurred: {str(exception)}")

def HashUsingMD5():
    HashPhoneNumbers(hashlib.md5)

def HashUsingSHA1():
    HashPhoneNumbers(hashlib.sha1)

def HashUsingSHA256():
    HashPhoneNumbers(hashlib.sha256)

def HashUsingSHA512():
    HashPhoneNumbers(hashlib.sha512)


root = tk.Tk()
root.title("Phone Number Hashing")

LabelInputFile = tk.Label(root, text="Select file with phone numbers:")
LabelInputFile.grid(row=0, column=0, padx=10, pady=10, sticky="w")
EntryInputFile = tk.Entry(root, width=50)
EntryInputFile.grid(row=0, column=1, padx=10, pady=10, sticky="ew")
ButtonSelectInputFile = tk.Button(root, text="Select", command=SelectInputFile)
ButtonSelectInputFile.grid(row=0, column=2, padx=10, pady=10)

LabelOutputFile = tk.Label(root, text="Select file to save hashes:")
LabelOutputFile.grid(row=1, column=0, padx=10, pady=10, sticky="w")
EntryOutputFile = tk.Entry(root, width=50)
EntryOutputFile.grid(row=1, column=1, padx=10, pady=10, sticky="ew")
ButtonSelectOutputFile = tk.Button(root, text="Select", command=SelectOutputFile)
ButtonSelectOutputFile.grid(row=1, column=2, padx=10, pady=10)

FrameHashButtons = tk.Frame(root)
FrameHashButtons.grid(row=2, column=0, columnspan=3, pady=10)

ButtonMD5 = tk.Button(FrameHashButtons, text="MD5", command=HashUsingMD5)
ButtonMD5.grid(row=0, column=0, padx=5, pady=5)
ButtonSHA1 = tk.Button(FrameHashButtons, text="SHA-1", command=HashUsingSHA1)
ButtonSHA1.grid(row=0, column=1, padx=5, pady=5)
ButtonSHA256 = tk.Button(FrameHashButtons, text="SHA-256", command=HashUsingSHA256)
ButtonSHA256.grid(row=0, column=2, padx=5, pady=5)
ButtonSHA512 = tk.Button(FrameHashButtons, text="SHA-512", command=HashUsingSHA512)
ButtonSHA512.grid(row=0, column=3, padx=5, pady=5)

root.grid_columnconfigure(0, weight=0)  # Label column
root.grid_columnconfigure(1, weight=1)  # Entry column (stretched more)
root.grid_columnconfigure(2, weight=0)

root.mainloop()
